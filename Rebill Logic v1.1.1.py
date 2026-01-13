
from openpyxl import Workbook
from collections import defaultdict

def export_to_excel_with_customer_names(audit_trails, ra_number, invoice_no, customer_1, customer_2, customer_1_name, customer_2_name, output_path=None):
    wb = Workbook()
    default_sheet = wb.active
    wb.remove(default_sheet)

    grouped_lines = defaultdict(list)
    for invoice, item_no, qty, price in audit_trails:
        grouped_lines[invoice].append((item_no, qty, price))

    for invoice_id, lines in grouped_lines.items():
        ws = wb.create_sheet(title=invoice_id)

        journal_headers = [
            'Type', 'No.', 'Description', 'Location Code', 'Quantity', 'Unit of Measure Code',
            'Unit Price Excl. Tax', 'Tax Group Code', 'Line Amount Excl. Tax', 'Amount Including Tax',
            'Line Discount %', 'Qty. to Assign', 'Qty. Assigned'
        ]
        ws.append(journal_headers)

        for item_no, qty, price in lines:
            qty_float = float(qty)
            price_float = float(price)
            line_amount = qty_float * price_float

            unit_price_fmt = f"{price_float:.2f}"
            line_amount_fmt = f"{line_amount:.2f}"

            ws.append([
                "G/L Account", 485300, "MISC. ALLOWANCES", "W01", qty_float, "EA",
                unit_price_fmt, "NONTAXABLE", line_amount_fmt, line_amount_fmt,
                " ", 0, " "
            ])
            description = f"({qty}) {item_no} @ ${unit_price_fmt} EA {invoice_id}"
            ws.append([
                " ", " ", description, " ", " ", " ",
                " ", " ", " ", 0,
                " ", 0, " "
            ])

        rebill_note = f"REBILL {invoice_id} {customer_1} TO {customer_2}"
        ws.append([" ", " ", rebill_note, " ", " ", " ", " ", " ", " ", 0, " ", 0, " "])
        ws.append([" ", " ", f"{customer_1_name} to {customer_2_name}", " ", " ", " ", " ", " ", " ", 0, " ", 0, " "])

    ws_sales = wb.create_sheet(title=f"Sales_{invoice_no}")

    sales_headers = [
        "Type", "No.", "Description", "Location Code", "Quantity", "Unit of Measure Code",
        "Unit Price Excl. Tax", "Tax Group Code", "Line Discount %", "Line Amount Excl. Tax",
        "Amount Including Tax", "Qty. to Assign"
    ]
    ws_sales.append(sales_headers)

    for invoice_id, lines in grouped_lines.items():
        for item_no, qty, price in lines:
            qty_float = float(qty)
            price_float = float(price)
            line_amount = qty_float * price_float

            unit_price_fmt = f"{price_float:.2f}"
            line_amount_fmt = f"{line_amount:.2f}"
            description = f"({qty}) {item_no} @ ${unit_price_fmt} EA {invoice_id}"

            ws_sales.append([
                "G/L Account", 485300, description, "W01", qty_float, "EA",
                unit_price_fmt, "", "", line_amount_fmt, line_amount_fmt, 0
            ])

    rebill_note = f"REBILL {invoice_no} {customer_1} TO {customer_2}"
    ws_sales.append([" ", " ", rebill_note, " ", " ", " ", " ", " ", " ", " ", " ", 0])
    ws_sales.append([" ", " ", f"{customer_1_name} to {customer_2_name}", " ", " ", " ", " ", " ", " ", " ", " ", 0])

    file_base = invoice_no.replace(" ", "").replace("&", "_")
    if not output_path:
        output_path = f"{file_base}_Rebill.xlsx"
    wb.save(output_path)
    return output_path
