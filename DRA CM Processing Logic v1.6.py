"""
DRA Credit Memo Processor
Version: 1.6
Last updated: 2025-09-22

What’s new (v1.6):
- Parenthetical quantity now renders as an integer when whole (e.g., (1) not (1.0)).

What’s new (v1.5):
- Output Excel filename is now exactly the RA number provided (e.g., RA109218.xlsx).
- Keeps our standard NAV import layout and values:
  G/L 488000 "DEFECTIVE ALLOWANCES", Location "W01", Tax Group "NONTAXABLE".
- Appends closing lines:
  "DAMAGES TO <RA#>" and "DESTROY IN FIELD".
- Adds a helper to generate the standard email draft text where the credit memo
  reference remains literally "SCR-000000" per SOP.
"""

from dataclasses import dataclass
from typing import List, Tuple
from openpyxl import Workbook


__version__ = "1.5"


@dataclass
class LineItem:
    invoice: str
    item_no: str
    qty: float
    unit_price: float


def build_export(audit_trails: List[Tuple[str, str, float, float]], ra_number: str) -> str:
    """
    Build an Excel file for NAV import using the SOP layout.
    Returns the absolute path to the saved file.
    audit_trails: list of tuples (invoice, item_no, qty, unit_price)
    ra_number: string like 'RA109218' which is also used as the output filename.
    """
    # Accept either tuples or LineItem dataclasses
    _rows: List[LineItem] = []
    for row in audit_trails:
        if isinstance(row, LineItem):
            _rows.append(row)
        else:
            invoice, item_no, qty, price = row
            _rows.append(LineItem(invoice=invoice, item_no=item_no, qty=float(qty), unit_price=float(price)))

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    headers = [
        'Type', 'No.', 'Description', 'Location Code', 'Quantity', 'Unit of Measure Code',
        'Unit Price Excl. Tax', 'Tax Group Code', 'Line Amount Excl. Tax', 'Amount Including Tax',
        'Line Discount %', 'Qty. to Assign', 'Qty. Assigned'
    ]
    ws.append(headers)

    for li in _rows:
        line_amount = float(li.qty) * float(li.unit_price)
        unit_price_fmt = f"{float(li.unit_price):.2f}"
        line_amount_fmt = f"{line_amount:.2f}"

        # CM line
        ws.append([
            "G/L Account", 488000, "DEFECTIVE ALLOWANCES", "W01", li.qty, "EA",
            unit_price_fmt, "NONTAXABLE", line_amount_fmt, line_amount_fmt,
            " ", 0, " "
        ])
        # Descriptive memo line
        invoice_tag = f" {li.invoice}" if str(li.invoice).strip() else ""
        description = f"({li.qty}) {li.item_no} @ ${unit_price_fmt} EA{invoice_tag}"
        ws.append([
            " ", " ", description, " ", " ", " ",
            " ", " ", " ", 0,
            " ", 0, " "
        ])

    # Closing lines per SOP
    ws.append([" ", " ", f"DAMAGES TO {ra_number}", " ", " ", " ", " ", " ", " ", 0, " ", 0, " "])
    ws.append([" ", " ", "DESTROY IN FIELD", " ", " ", " ", " ", " ", " ", 0, " ", 0, " "])

    out_path = f"/mnt/data/{ra_number}.xlsx"
    wb.save(out_path)
    # Normalize any parenthetical quantity from decimals to integers, e.g., (1.0) -> (1)
    from openpyxl import load_workbook
    import re
    _wb = load_workbook(out_path)
    _ws = _wb.active
    _pat = re.compile(r"\((\d+)(?:\.0+)?\)")
    for _row in _ws.iter_rows(values_only=False):
        for _cell in _row:
            if isinstance(_cell.value, str) and '(' in _cell.value and ')' in _cell.value:
                _new = _pat.sub(lambda m: f"({int(m.group(1))})", _cell.value)
                if _new != _cell.value:
                    _cell.value = _new
    _wb.save(out_path)
    return out_path


def generate_email_draft(ra_number: str) -> str:
    """
    Returns the standard email draft. The credit memo reference remains
    literally 'SCR-000000' as per SOP.
    """
    return (
        "Hi,\n\n"
        f"Please see the attached credit memo SCR-000000 for {ra_number}.\n\n"
        "Best,"
    )


if __name__ == "__main__":
    # Example usage
    sample_lines = [
        # (invoice, item_no, qty, unit_price)
        ("SI522429", "PRAIR701LFS", 1, 99.99),
        ("SI534865", "CFF701MFS", 1, 14.99),
    ]
    ra = "RA109218"
    path = build_export(sample_lines, ra)
    print(f"Export written to: {path}")
    print("\nEmail draft:\n")
    print(generate_email_draft(ra))
